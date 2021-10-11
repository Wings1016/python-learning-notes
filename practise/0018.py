#!/usr/bin/env python3
# -*- coding: utf-8 -*-

'''
将 第 0015 题中的 city.xls 文件中的内容写到 city.xml 文件中，如下 所示：
    <?xmlversion="1.0" encoding="UTF-8"?>
    <root>
    <cities>
    <!-- 
    	城市信息
    -->
    {
    	"1" : "上海",
    	"2" : "北京",
    	"3" : "成都"
    }
    </cities>
    </root>
'''

from openpyxl import load_workbook
from xml.dom import minidom

def getdatafromexcel():
    wb = load_workbook('./city.xlsx')
    sheet = wb['Sheet']
    max_row = sheet.max_row     # 行数
    max_col = sheet.max_column  # 列数
    values = list()
    for i in range(max_row):
        values.append([s.value for s in sheet[i+1]])
    return values

def write2xml(data):
    #在内存中创建一个空的文档
    doc = minidom.Document() 
    #创建一个根节点Managers对象
    root = doc.createElement('root') 
    doc.appendChild(root)
    cities = doc.createElement('cities')
    data1 = '   城市信息\n'
    cities.appendChild(doc.createComment(data1))
    data2 = '{\n   “%s”:“%s”,\n   “%s”:”%s“,\n   “%s”:”%s“\n}' % (data[0][0],data[0][1],data[1][0],data[1][1],data[2][0],data[2][1])
    text = doc.createTextNode(data2)
    cities.appendChild(text)
    root.appendChild(cities)
    with open('./city.xml','w') as f:
        doc.writexml(f,indent='', addindent='', newl='\n', encoding="utf-8")


if __name__=='__main__':
    write2xml(getdatafromexcel())