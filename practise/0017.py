#!/usr/bin/env python3
# -*- coding: utf-8 -*-

'''
将 第 0014 题中的 student.xls 文件中的内容写到 student.xml 文件中，如下所示：

<?xml version="1.0" encoding="UTF-8"?>
<root>
<students>
<!-- 
	学生信息表
	"id" : [名字, 数学, 语文, 英文]
-->
{
	"1" : ["张三", 150, 120, 100],
	"2" : ["李四", 90, 99, 95],
	"3" : ["王五", 60, 66, 68]
}
</students>
</root>
'''

from openpyxl import load_workbook
from xml.dom import minidom

def getdatafromexcel():
    wb = load_workbook('./student.xlsx')
    sheet = wb['Sheet']
    max_row = sheet.max_row     # 行数
    max_col = sheet.max_column  # 列数
    values = list()
    for i in range(max_row):
        values.append([s.value for s in sheet[i+1]])
    return values

def write2xml(data):
    #在内存中创建一个空的文档
    doc = minidom.Document() 
    #创建一个根节点Managers对象
    root = doc.createElement('root') 
    doc.appendChild(root)
    student = doc.createElement('student')
    data1 = '   学生信息表\n    "id" : [名字, 数学, 语文, 英文]\n'
    student.appendChild(doc.createComment(data1))
    data2 = '{\n   %s,\n   %s,\n   %s\n}' % (data[0],data[1],data[2])
    text = doc.createTextNode(data2)
    student.appendChild(text)
    root.appendChild(student)
    with open('./student.xml','w') as f:
        doc.writexml(f,indent='', addindent='', newl='\n', encoding="utf-8")


if __name__=='__main__':
    write2xml(getdatafromexcel())