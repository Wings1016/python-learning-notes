#!/usr/bin/env python3
# -*- coding: utf-8 -*-

'''
将 第 0016 题中的 numbers.xls 文件中的内容写到 numbers.xml 文件中，如下所示：
<?xml version="1.0" encoding="UTF-8"?>
<root>
<numbers>
<!-- 
	数字信息
-->

[
	[1, 82, 65535],
	[20, 90, 13],
	[26, 809, 1024]
]

</numbers>
</root>
'''

from openpyxl import load_workbook
from xml.dom import minidom

def getdatafromexcel():
    wb = load_workbook('./numbers.xlsx')
    sheet = wb['Sheet']
    max_row = sheet.max_row     # 行数
    max_col = sheet.max_column  # 列数
    values = list()
    for i in range(max_row):
        values.append([s.value for s in sheet[i+1]])
    return values

def write2xml(data):
    #在内存中创建一个空的文档
    doc = minidom.Document() 
    #创建一个根节点Managers对象
    root = doc.createElement('root') 
    doc.appendChild(root)
    numbers = doc.createElement('numbers')
    data1 = '\n     数字信息\n'
    numbers.appendChild(doc.createComment(data1))
    data2 = '\n[\n   %s,\n   %s,\n   %s\n]\n' % (data[0],data[1],data[2])
    text = doc.createTextNode(data2)
    numbers.appendChild(text)
    root.appendChild(numbers)
    with open('./numbers.xml','w') as f:
        doc.writexml(f,indent='', addindent='', newl='\n', encoding="utf-8")


if __name__=='__main__':
    write2xml(getdatafromexcel())