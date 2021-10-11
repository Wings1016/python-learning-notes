#!/usr/bin/env python3
# -*- coding: utf-8 -*-

'''
class Solution:
    def twoSum(self, nums, target: int):
        result = []
        nums.sort()
        for i in range(len(nums)):
            for j in range(len(nums)):
                if i<j:
                    result[i][j]=nums[i]+nums[j]
                    if result[i][j]==target:
                        return [i,j]

a=[2,7,11,15]
so = Solution()
result = so.twoSum(a,9)
print(result)


import numpy as np

a=np.empty((2,3))
x=np.linspace(0,2*np.pi,5)
f = np.sin(x)

def f1(x,y):
    return 10*x+y

b = np.fromfunction(f1,(5,4),dtype=int)
print(b)

print(np.random.random((3,4)))


matrix = [
    [1, 2, 3, 4],
    [5, 6, 7, 8],
    [9, 10, 11, 12],
]

print([col[0] for col in matrix])
'''

'''
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
wb = Workbook()
dest_filename = 'empty_book.xlsx'
ws1 = wb.active
ws1.title = "range names"
for row in range(1, 40):
    ws1.append(range(600))
ws2 = wb.create_sheet(title="Pi")
ws2['F5'] = 3.14
ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)
wb.save(filename = dest_filename)
'''

'''
class Bag:
    def __init__(self):
        self.data = []

    def add(self, x):
        self.data.append(x)

    def addtwice(self, x):
        self.add(x)
        self.add(x)

bag = Bag()
bag.add(1)
bag.addtwice(2)
print(bag.data) # [1,2,2]
'''

'''
import struct

with open('/Users/herry/Downloads/cert.zip', 'rb') as f:
    data = f.read()

start = 0
for i in range(3):                      # show the first 3 file headers
    start += 14
    fields = struct.unpack('<IIIHH', data[start:start+16])
    crc32, comp_size, uncomp_size, filenamesize, extra_size = fields

    start += 16
    filename = data[start:start+filenamesize]
    start += filenamesize
    extra = data[start:start+extra_size]
    print(filename, hex(crc32), comp_size, uncomp_size)

    start += extra_size + comp_size     # skip to the next header
'''

from heapq import heapify, heappop, heappush
# heapq提供了基于常规列表来实现堆的函数。 最小值的条目总是保持在位置零
data = [1, 3, 5, 7, 9, 2, 4, 6, 8, 0]
heapify(data)                      # rearrange the list into heap order
print(data)
heappush(data, 22)                 # add a new entry
print(data)
print([heappop(data) for i in range(3)])  # fetch the three smallest entries